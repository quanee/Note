import xlrd

import sqlite3

file = "H:\\xls\\全国省市县列表.xls"

data = xlrd.open_workbook(file)

table = data.sheets()[0]  # 第一个sheets

datalist = []

for i in range(1, table.nrows):  # 总行数

    datalist.append(tuple(table.row_values(i)))

conn = sqlite3.connect('d:\\database\\country.db')  # 数据库文件的路径

cursor = conn.cursor()

cursor.execute("create table country(province varchar(20),num int,town varchar(30),town varchar(30))")  # #创建表

sql = 'insert into country(province,num,city,town) values(?,?,?,?)'

cursor.executemany(sql, datalist)  # 插入数据

conn.commit()  # 提交数据到数据库

conn.close()  # 关闭连接

print("导入完成")
# 将excel导入mysql
#/usr/bin/env python3

# -*- coding:utf-8 -*-

import os
import sys
import getopt
import openpyxl
import pymysql
import datetime
import re
import itertools

# 数据库连接配置文件

__db_config = {
    'host': '127.0.0.1',
    'port': 3306,
    'user': 'your.db.username',
    'password': 'your.db.password',
    'db': 'xls',
    'charset': 'utf8'
}

####################


def usage():

    print('xls2db.py usage:')

    print(' xls2db filename(xlsx)')

    print(' version: 1.0.1')

    print(' by sun')

    exit(1)

# 主函数执行


if __name__ == '__main__':

    if len(sys.argv) != 2:

        usage()

fname = sys.argv[1]

if not os.path.exists(fname):

    print("file not exist:", fname)

exit(2)

# 创建链接

connection = pymysql.connect(**__db_config)

print("loadfile:", fname)

wb = openpyxl.load_workbook(fname)

for shname in wb.get_sheet_names():

    ws = wb.get_sheet_by_name(shname)

    # 过滤空表

    if ws.max_row <= 1:

        break

# 构建建表SQL

sql = 'DROP TABLE IF EXISTS `xls`.`%s`;\n' % (shname)

sql += 'CREATE TABLE `xls`.`%s` (\n `indx` INT NOT NULL AUTO_INCREMENT,\n' % (shname)

# 构建表的结构

for col in ws.iter_cols(max_row=2):

    if not isinstance(col[0].value, str):

        break

if col[1].is_date:

    sql += ' `%s` DATETIME NULL DEFAULT NULL, \n' % (col[0].value)

else:

    sql += ' `%s` VARCHAR(64) NULL DEFAULT NULL,\n' % (col[0].value)

    sql += ' PRIMARY KEY (`indx`) \n) DEFAULT CHARACTER SET = utf8;\n'

with connection.cursor() as cursor:

    row_count = cursor.execute(sql)

print('create table:', shname)

# 构建自增主键

indx = itertools.count(1, 1)

# 构建插入语句

sql = 'INSERT IGNORE INTO `xls`.`%s` ' % (shname)